import streamlit as st
import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import tempfile
import io

# =====================================================================================
# FUNÇÃO UNIFICADA: PROCESSA O ARQUIVO INDEPENDENTE DO TIPO
# =====================================================================================
def processar_arquivo_generico(caminho_csv):
    """
    Lê um arquivo CSV e tenta identificar automaticamente se é PRODUTO ACABADO
    ou BOBINA baseando-se na estrutura dos dados da coluna 4.
    """
    try:
        # Tenta ler com utf-8, se falhar vai de latin1
        try:
            df = pd.read_csv(caminho_csv, header=None, encoding='utf-8', dtype=str)
        except UnicodeDecodeError:
            df = pd.read_csv(caminho_csv, header=None, encoding='latin1', dtype=str)
    except Exception as e:
        return None, f"Erro ao ler arquivo: {e}"

    dados_processados = []
    
    # Verifica se o arquivo está vazio ou muito curto
    if df.empty:
        return pd.DataFrame(), None

    # Iterar linha a linha para processar
    for index, row in df.iterrows():
        # Pula linhas quebradas
        if len(row) < 5:
            continue

        # Dados brutos básicos
        dt_leitura = str(row[0]).strip()
        hr_leitura = str(row[1]).strip()
        coluna_tipo = str(row[3]).strip() # Geralmente Code128, QR_CODE ou vazio no PA
        dados_lidos = str(row[4]).strip() # A string principal com os dados

        # Filtro básico: se não tiver data, ignora (sujeira de leitura)
        if "date" in dt_leitura.lower() or not dt_leitura[0].isdigit():
            continue

        # --- ESTRUTURA DO DICIONÁRIO PADRÃO (Super Conjunto de Colunas) ---
        nova_linha = {
            "Data da Leitura": dt_leitura,
            "Hora da Leitura": hr_leitura,
            "Filial": None,
            "Código": None,
            "Armazém": None,
            "Lote": None,
            "Peso": None,
            "Localização": os.path.splitext(os.path.basename(caminho_csv))[0]
        }

        # ====================================================================
        # TESTE 1: É PRODUTO ACABADO? (Padrão: "XXX-XXX - YYY")
        # ====================================================================
        if " -" in dados_lidos:
            try:
                # Lógica de fatiar (Split) do Produto Acabado
                partes_maiores = dados_lidos.split(" -", 1) # Divide no " -"
                
                # Lado Esquerdo (Filial-Codigo)
                parte_esq = partes_maiores[0].split("-")
                filial = parte_esq[0].strip() if len(parte_esq) > 0 else ""
                codigo = parte_esq[1].strip() if len(parte_esq) > 1 else ""
                
                # Lado Direito (Armazem-Lote-Peso...)
                parte_dir = partes_maiores[1].split("-") if len(partes_maiores) > 1 else []
                
                # Mapeamento posicional (ajuste conforme seu padrão de Prod Acabado)
                # Padrão esperado na direita: Armazem - Lote - Peso ...
                armazem = parte_dir[0].strip() if len(parte_dir) > 0 else ""
                lote = parte_dir[1].strip() if len(parte_dir) > 1 else ""
                peso_str = parte_dir[2].strip() if len(parte_dir) > 2 else "0"
                
                # Tratamento do Peso (Prod Acabado divide por 1000)
                try:
                    peso_val = float(peso_str) / 1000.0
                except:
                    peso_val = 0.0

                nova_linha["Filial"] = filial
                nova_linha["Código"] = codigo
                nova_linha["Armazém"] = armazem
                nova_linha["Lote"] = lote
                nova_linha["Peso"] = peso_val
                
                dados_processados.append(nova_linha)
                continue # Linha processada com sucesso como PA, vai para a próxima
            
            except Exception:
                # Se falhar o split do PA, cai para a tentativa de Bobina abaixo
                pass

        # ====================================================================
        # TESTE 2: É BOBINA? (Logica "Detetive")
        # ====================================================================
        
        # Formata data para ficar bonito (dd/mm/yyyy) se possível
        try:
            nova_linha["Data da Leitura"] = datetime.strptime(dt_leitura, '%m-%d-%Y').strftime('%d/%m/%Y')
        except:
            pass # Mantém original se falhar

        lote_b = "erro"
        peso_b = 0.0

        # Lógica CODE128 (Asteriscos)
        if coluna_tipo == 'Code128' or '*' in dados_lidos:
            if ' ' in dados_lidos:
                 lote_b, peso_b = "erro de leitura", 0
            elif '*' in dados_lidos:
                try:
                    partes = dados_lidos.split('*')
                    # Tenta adivinhar onde está o peso e o lote
                    if dados_lidos.startswith('*'): 
                        lote_b = partes[3].strip()
                        peso_b = float(partes[2].strip()) / 1000.0
                    else: 
                        lote_b = partes[2].strip()
                        peso_b = float(partes[1].strip()) / 1000.0
                except:
                    lote_b, peso_b = "erro Code128/*", 0
            elif dados_lidos.isdigit() and len(dados_lidos) <= 5:
                 peso_b, lote_b = float(dados_lidos)/1000.0, ""
            else:
                 lote_b, peso_b = dados_lidos, 0

        # Lógica QR CODE / DATAMATRIX (JSON ou Texto Complexo)
        elif coluna_tipo in ['QR_CODE', 'QR', 'CODE_39', 'CODE_128'] or '{' in dados_lidos:
            # 1. JSON
            if '{' in dados_lidos and '}' in dados_lidos:
                try:
                    partes = dados_lidos.split('{', 1)
                    identificador = partes[0].strip('"-')
                    dados_json = json.loads('{' + partes[1])
                    peso_b = float(dados_json.get('peso', 0))
                    lote_b = identificador
                except:
                    lote_b = "erro QR/JSON"

            # 2. Novo Formato (Vírgula para peso)
            elif ',' in dados_lidos and '-' in dados_lidos:
                try:
                    partes_virgula = dados_lidos.split(',')
                    if len(partes_virgula) > 1 and partes_virgula[-1].replace('.', '', 1).isdigit():
                        peso_str = partes_virgula[-1].strip()
                        parte_lote_completa = ','.join(partes_virgula[:-1])
                        partes_hifen = parte_lote_completa.split('-')
                        lote_b = partes_hifen[-2].strip()
                        peso_completo_str = f"{partes_hifen[-1].strip()},{peso_str}"
                        peso_b = float(peso_completo_str.replace(',', '.'))
                        # Nota: Não divide por 1000 aqui, conforme logica original da Bobina
                    else:
                        raise ValueError("Formato virgula invalido")
                except:
                    # Fallback para formato antigo
                    try:
                        partes = dados_lidos.split('-')
                        lote_b = partes[3].strip()
                        peso_b = float(partes[-1].strip()) / 1000.0
                    except:
                        lote_b = "erro QR/Formato"
            
            # 3. Formato Simples/Antigo (Só hifens)
            else:
                 try:
                     partes = dados_lidos.split('-')
                     # Assume que o último é peso e o antepenultimo ou especifico é lote
                     if len(partes) >= 4:
                        lote_b = partes[3].strip() # Posição comum em etiquetas antigas
                        peso_b = float(partes[-1].strip()) / 1000.0
                     else:
                        lote_b = dados_lidos
                        peso_b = 0
                 except:
                     lote_b = dados_lidos
                     peso_b = 0

        else:
            # Se não reconheceu nada, joga o dado cru no Lote
            lote_b = dados_lidos
        
        nova_linha["Lote"] = lote_b
        nova_linha["Peso"] = peso_b
        dados_processados.append(nova_linha)

    return pd.DataFrame(dados_processados), None

# =====================================================================================
# INTERFACE DO STREAMLIT (UI)
# =====================================================================================

st.set_page_config(page_title="Conversor de Inventário Dox", layout="wide")
st.title("Conversor de Inventário Unificado")
st.markdown("---")

# --- INPUTS DO USUÁRIO ---
col1, col2 = st.columns([2, 1])
with col1:
    uploaded_files = st.file_uploader(
        "Importar arquivos .csv (Aceita Bobina e Produto Acabado misturados)",
        type="csv",
        accept_multiple_files=True
    )

with col2:
    st.info("Configurações de Saída")
    nome_arquivo_usuario = st.text_input("Nome do Arquivo Final (sem .xlsx):", value="")

# --- BOTÃO E LÓGICA DE EXECUÇÃO ---
if st.button("Converter Arquivos", type="primary"):
    if not uploaded_files:
        st.warning("⚠️ Por favor, carregue pelo menos um arquivo .csv.")
    else:
        with st.spinner("Processando... O sistema está identificando o tipo de cada arquivo."):
            try:
                todos_dfs = []
                
                # Processa cada arquivo individualmente na memória
                for uploaded_file in uploaded_files:
                    # Salva temporariamente para processar
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_file:
                        tmp_file.write(uploaded_file.getbuffer())
                        tmp_path = tmp_file.name
                    
                    # Chama a função unificada
                    df_temp, erro = processar_arquivo_generico(tmp_path)
                    
                    if erro:
                        st.error(f"Erro no arquivo {uploaded_file.name}: {erro}")
                    elif not df_temp.empty:
                        # Restaura o nome original para a coluna localização (pois o tempfile tem nome aleatorio)
                        df_temp["Localização"] = uploaded_file.name.replace('.csv', '')
                        todos_dfs.append(df_temp)
                    
                    # Limpa arquivo temporário
                    os.unlink(tmp_path)

                if todos_dfs:
                    # Consolida tudo
                    df_final = pd.concat(todos_dfs, ignore_index=True)
                    
                    # Tratamento final de estética (Zeros a esquerda no Armazem, etc)
                    if "Armazém" in df_final.columns:
                        df_final["Armazém"] = df_final["Armazém"].fillna('').apply(lambda x: str(x).split('.')[0].zfill(2) if str(x).replace('.','').isdigit() else str(x))
                    
                    # Gera o Excel em memória
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_final.to_excel(writer, index=False, sheet_name='Inventario Geral')
                        ws = writer.sheets['Inventario Geral']
                        
                        # Formatação Visual (Largura e Casas Decimais)
                        for col in ws.columns:
                            max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
                            
                            # Se for a coluna de Peso (assume que peso está na coluna G=7 ou H, busca pelo cabeçalho)
                            if col[0].value == "Peso":
                                for cell in col[1:]: # Pula cabeçalho
                                     cell.number_format = '0.000'

                    output.seek(0)
                    
                    # Define o nome do arquivo
                    if nome_arquivo_usuario.strip():
                        nome_download = f"{nome_arquivo_usuario.strip()}.xlsx"
                    else:
                        nome_download = "Inventario.xlsx"

                    st.success(f"✅ Sucesso! {len(todos_dfs)} arquivos processados.")
                    
                    st.download_button(
                        label="📥 Baixar Excel Consolidado",
                        data=output,
                        file_name=nome_download,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Os arquivos foram lidos, mas nenhum dado válido foi encontrado.")

            except Exception as e:
                st.error(f"Ocorreu um erro crítico: {e}")